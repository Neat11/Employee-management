from openpyxl import load_workbook
import datetime
import win32com.client as win32
from destination import path
wb = load_workbook(path)
sheet = wb["MAST1"]

def get_maximum_rows(sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows

def get_maximum_cols(sheet):
        max_col = 0
        for i in range(1, sheet.max_column+2):
            if (sheet.cell(row=2, column= i).value == None) and (sheet.cell(row=1, column= i).value == None):
                max_col = i
                break
        return max_col

rows = get_maximum_rows(sheet)
cols = get_maximum_cols(sheet)

def rem_emp(tfcode, name, date_left, keylist):
    global rows,cols,wb,sheet
    dts = date_left.split('-')
    print(dts)
    for j in range(rows-1):
        if sheet[chr(ord('A')+keylist.index("TFCODE"))+str(j+2)].value == tfcode:
            if sheet[chr(ord('A')+keylist.index("NAME"))+str(j+2)].value == name:
                try: 
                    datetime.datetime(int(dts[2]),int(dts[1]),int(dts[0]))
                    sheet[chr(ord('A')+keylist.index("DATELEFT"))+str(j+2)] = date_left
                    sheet[chr(ord('A')+keylist.index("STATUS"))+str(j+2)] = "Inactive"
                    wb.save(path)
                    return "employee removed successfully"
                except:
                    return "Invalid Date"
            else:
                return "Name does not match tfcode"
    else:
        return "No tfcode found"


