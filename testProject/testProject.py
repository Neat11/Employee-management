from datetime import datetime
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.properties import StringProperty,NumericProperty
from openpyxl import Workbook, load_workbook
import add_employee_logic as ael
import rem_employee_logic as rel
import advance_pay as ap
import salary_logic as sl
import Attendance_logic as Attend
import Print_logic as pl
from destination import path
class findDimensions:
    def get_maximum_cols(sheet):
        max_col = 0
        for i in range(1, sheet.max_column+2):
            if (sheet.cell(row=2, column= i).value == None) and (sheet.cell(row=1, column= i).value == None):
                max_col = i-1
                break
        return max_col

    def get_maximum_rows(sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows

class Password(Screen):
    password = StringProperty("123")


class Menu(Screen):
    title = StringProperty('Company Name')
    w = NumericProperty(0.3)
    h = NumericProperty(0.45)


class EmployeeAdd(Screen):
    keyList = []
    headings =[]
    heading_new_values={}
    error_msg = StringProperty("")
    try:
        wb = load_workbook(path)
        sheet = wb["MAST1"]
    except:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "MAST1"
        wb.create_sheet("M1",1)
        wb.create_sheet("MNTHSUM1",2)
    
    cols = findDimensions.get_maximum_cols(sheet)
    if cols>0:
        char = 'A'
        for i in range(cols):
            headings.append(sheet[char+'1'].value)
            heading_new_values[sheet[char+'1'].value] = ""
            char = chr(ord(char)+1)
    else:
        headings = ["NAME","DTJOIN","LEAVE_INIT","LALLOW","LAVAIL","BPAY","T_ALL","SW_ALL","HRA","M_ALL","O_ALL","ADV_BAL","ADV_DEDRT","REF","TFCODE","LVMEBAL","COMP","PTAXS","DEPOSIT","DATELEFT","STATUS"]
        heading_new_values = {i : "" for i in headings}
        mnth_headings=["TFCODE","NAME","MONTH","ABSENTS	LVDAYS","PAYDAYS","BPAY","T_ALL","SW_ALL","HRA","MED_ALL","O_ALL","ADV_DED","NAMT","LVBAL","COMP","PTAX","DEPOSIT"]
        m1_headings = ["DATE","NAME","INTIME","OUTTIME","LDESC","LCHECK"]
        for i in range(len(headings)):
            sheet.cell(1,i+1).value = headings[i]
        for i in range(len(mnth_headings)):
            wb["MNTHSUM1"].cell(1,i+1).value = mnth_headings[i]
        for i in range(len(m1_headings)):
            wb['M1'].cell(1,i+1).value = m1_headings[i]
    wb.save(path)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        main_layout = BoxLayout(orientation="vertical")
        self.add_widget(main_layout)
        grid_layout = GridLayout(cols = 2)
        main_layout.add_widget(grid_layout)
        for i in self.headings[:-3]:
            c = AnchorLayout()
            grid_layout.add_widget(c)
            b = BoxLayout(size_hint=(0.7,0.6))
            l = Label(size_hint=(0.4,1),text=str(i)+":", halign="center")
            self.ids[i+"_L"] = l
            t = TextInput(text="", multiline=False,write_tab=False)
            self.ids[i+"_T"] = t 
            b.add_widget(l)
            b.add_widget(t)
            c.add_widget(b)
        err = Label(text = "", color = (1,0,0,1))
        self.ids["err"] = err
        grid_layout.add_widget(err)
        anchor_layout = AnchorLayout(anchor_x="right", size_hint=(1,None), height=80)
        b = BoxLayout(size_hint=(None,None), height=70,width=300, padding=20, spacing = 20)
        submit = Button(text="Submit", on_release=self.submit_button)
        back = Button(text="Back", on_release=self.changer)
        b.add_widget(submit)
        b.add_widget(back)
        anchor_layout.add_widget(b)
        main_layout.add_widget(anchor_layout)

    
    def changer(self, *args):
        self.manager.current = 'menu'

    def on_size(self, instance, value ):
        self.font_size=0.4*self.width
        for i in self.headings[:-3]:
            self.ids[i+"_L"].font_size = 0.02*self.width
        self.ids["err"].font_size = 0.02*self.width

    def submit_button(self, values):
        for i in self.headings[:-3]:
            self.heading_new_values[i] = self.ids[i+"_T"].text
        self.heading_new_values["STATUS"] = "Active"
        err = ael.EmployeeAdd.add_emp(ael.EmployeeAdd, self.heading_new_values)
        if err == "Employee added successfully":
            for i in self.headings[:-3]:
                self.ids[i+"_T"].text = ""
        self.ids["err"].text = err
        

class EmployeeRem(Screen):

    err = StringProperty("")
    wb = load_workbook(path)
    sheet = wb["MAST1"]
    keyList = []

    cols = findDimensions.get_maximum_cols(sheet)
    if cols>0:
        char = 'A'
        for i in range(cols):
            keyList.append(sheet[char+'1'].value)
            char = chr(ord(char)+1)

    def submit(self):
        self.err = rel.rem_emp(self.ids["tfcode3"].text,self.ids["name3"].text,self.ids["date3"].text,self.keyList)


class Salary(Screen):
    wb = load_workbook(path)
    sheet = wb['MAST1']
    keyList = []
    mnthsum = wb['MNTHSUM1']
    rows_m = findDimensions.get_maximum_rows(mnthsum)
    cols = findDimensions.get_maximum_cols(sheet)
    if cols>0:
        char = 'A'
        for i in range(cols):
            keyList.append(sheet[char+'1'].value)
            char = chr(ord(char)+1)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.main_layout = AnchorLayout()
        self.anchor_layout = AnchorLayout(anchor_x="right", anchor_y="bottom", padding=(20,20))
        self.add_widget(self.main_layout)
        self.boxlayout = BoxLayout(size_hint=(0.4,0.1),orientation = "vertical")
        self.boxlayout.add_widget(Label(text = "Today's Date:", font_size=0.2*self.width,color=(0,0,0,1)))
        b = Button(text = "Back", on_release=self.changer, size_hint=(None,None),height=30,width=120)
        date = TextInput(text="", halign='center', on_text_validate=self.switch, multiline=False)
        self.ids["DOA"] = date
        self.boxlayout.add_widget(date)
        self.main_layout.add_widget(self.boxlayout)
        self.anchor_layout.add_widget(b)
        self.main_layout.add_widget(self.anchor_layout)

    def changer(self, values):
        self.manager.current = 'menu'

    def switch(self, dt):
        try: 
            dates = self.ids["DOA"].text.split(".")
            datetime(int(dates[2]),int(dates[1]), int(dates[0]))
            rows = findDimensions.get_maximum_rows(self.sheet)
            grid = GridLayout(cols=3,size_hint=(1,None))
            grid.add_widget(Label(text="Names",halign='center',size_hint_y=None,height=0.04*self.width, font_size=0.02*self.width,outline_width=2))
            grid.add_widget(Label(text="Dates",halign='center',size_hint_y=None,height=0.04*self.width, font_size=0.02*self.width,outline_width=2))
            grid.add_widget(Label(text="Salary",halign='center',size_hint_y=None,height=0.04*self.width, font_size=0.02*self.width,outline_width=2))
            for i in range(2,rows+1):
                if self.sheet.cell(row = i, column=(self.keyList.index("STATUS")+1)).value=="Active":
                    grid.add_widget(Label(text=self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value,outline_width=2, font_size=0.02*self.width,size_hint_y=None,height=0.04*self.width))
                    t = TextInput(text="",hint_text="dd.mm.yyyy-\"\",...",size_hint_y=None,height=0.04*self.width, font_size=0.02*self.width,multiline=False)
                    self.ids[self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value+'_TS'] = t
                    l = Label(text="",size_hint_y=None,height=0.04*self.width, font_size=0.02*self.width,outline_width=2)
                    self.ids[self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value+'_LS'] = l
                    grid.add_widget(t)
                    grid.add_widget(l)
            self.main_layout.remove_widget(self.boxlayout)
            self.main_layout.remove_widget(self.anchor_layout)
            grid.bind(minimum_height=grid.setter('height'))
            sv = ScrollView(size_hint=(0.9,0.8), do_scroll_x=False)
            sv.add_widget(grid)
            self.main_layout.add_widget(sv)
            ax = AnchorLayout(anchor_x='center', anchor_y='bottom', size_hint=(1,1))
            self.str_d1 = self.mnthsum.cell(row = self.rows_m,column=3).value
            self.str_d2 = datetime.strptime(self.ids["DOA"].text, "%d.%m.%Y")
            self.delta = self.str_d2-self.str_d1
            datem = (str(self.mnthsum.cell(row = self.rows_m,column=3).value).split(' ')[0]).split('-')
            lb = Label(text=f"{datem[2]}.{datem[1]}.{datem[0]} - {self.ids['DOA'].text} ({self.delta.days})", outline_width=2,halign='left')
            bx = BoxLayout(size_hint=(1,0.1),padding=20)
            but = Button(text='Calculate', on_release=self.salary_calculate,size_hint=(None,None),width = 100,height=40)
            self.ids["Calculate_B"] = but
            bx.add_widget(lb)
            bx.add_widget(but)
            ax.add_widget(bx)
            self.main_layout.add_widget(ax)
        except Exception as e:
            print(e)

    def salary_calculate(self, dt):
        if self.ids["Calculate_B"].text != "Calculate":
            self.print()
        mast1_headings = self.keyList
        mast1_rows=findDimensions.get_maximum_rows(self.sheet)
        values={}
        for i in range(2,mast1_rows+1):
             if self.sheet.cell(row = i, column=(self.keyList.index("STATUS")+1)).value == "Active":
                values[self.sheet.cell(row = i, column=(self.keyList.index("TFCODE")+1)).value] = self.ids[self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value+'_TS'].text 
        vals = sl.Salary_calculation.calculate(values,mast1_headings,self.str_d1, self.str_d2, self.delta.days)
        for i in range(2,mast1_rows+1):
            if self.sheet.cell(row = i, column=(self.keyList.index("STATUS")+1)).value == "Active":
                print(vals[self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value])
                self.ids[self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value+'_LS'].text = str(vals[self.sheet.cell(row = i, column=(self.keyList.index("NAME")+1)).value])
        self.ids["Calculate_B"].text= "Print"
    
    def print(self):
        pl.print_pls.make_file(self.str_d1, self.str_d2)
        

class Advance(Screen):
    err = StringProperty("")
    wb = load_workbook(path)
    sheet = wb["MAST1"]
    keyList = []
    

    cols = findDimensions.get_maximum_cols(sheet)
    if cols>0:
        char = 'A'
        for i in range(cols):
            keyList.append(sheet[char+'1'].value)
            char = chr(ord(char)+1)
    def submit(self):
        self.err = ap.add_advance(self.ids["tfcode4"].text,self.ids["name4"].text,self.ids["amount"].text,self.ids["months"].text,self.keyList)
        if self.err == "Advance Added":
            self.ids["tfcode4"].text,self.ids["name4"].text,self.ids["amount"].text,self.ids["months"].text = "","","","-"


class  Attendance(Screen):
    wb = load_workbook(path)
    sheet = wb["M1"]
    keyList = []

    cols = findDimensions.get_maximum_cols(sheet)
    if cols>0:
        char = 'A'
        for i in range(cols):
            keyList.append(sheet[char+'1'].value)
            char = chr(ord(char)+1)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        main_layout = BoxLayout(orientation="vertical")
        self.add_widget(main_layout)
        grid_layout = GridLayout(cols = 2)
        main_layout.add_widget(grid_layout)
        for i in self.keyList[:-2]:
            c = AnchorLayout()
            grid_layout.add_widget(c)
            b = BoxLayout(size_hint=(0.7,None), height=30)
            l = Label(size_hint=(0.4,1),text=str(i)+":", halign="center", color=(0,0,0,1))
            self.ids[i+"_L"] = l
            t = TextInput(text="", multiline=False,write_tab=False)
            self.ids[i+"_T"] = t 
            b.add_widget(l)
            b.add_widget(t)
            c.add_widget(b)
        err = Label(text = "", color = (1,0,0,1))
        self.ids["err"] = err
        grid_layout.add_widget(err)
        anchor_layout = AnchorLayout(anchor_x="right", size_hint=(1,None), height=80)
        b = BoxLayout(size_hint=(None,None), height=70,width=300, padding=20, spacing = 20)
        submit = Button(text="Submit", on_release=self.submit_button)
        back = Button(text="Back", on_release=self.changer)
        b.add_widget(submit)
        b.add_widget(back)
        anchor_layout.add_widget(b)
        main_layout.add_widget(anchor_layout)

    def submit_button(self, values):
        Attend.Att.mark_present(Attend.Att,self.ids[self.keyList[0]+"_T"].text,self.ids[self.keyList[1]+"_T"].text,self.ids[self.keyList[2]+"_T"].text,self.ids[self.keyList[3]+"_T"].text)
        self.ids["err"].text = "Data Recorded"
        for i in self.keyList[:-2]:
            self.ids[i+"_T"].text = "" 
     
    def changer(self, values):
        self.manager.current= 'menu'

class MyScreenManager(ScreenManager):
    def changescreen(self, value):
        self.current = value 


class TestApp(App):

    def build(self):
        self.sm = MyScreenManager()
        return self.sm


TestApp().run()
