from openpyxl import load_workbook
from destination import path
class Att:
    wb = load_workbook(path)
    m1 = wb["M1"]

    def get_maximum_cols(sheet):
        max_col = 0
        for i in range(1, sheet.max_column+2):
            if (sheet.cell(row=2, column= i).value == None) and (sheet.cell(row=1, column= i).value == None):
                max_col = i-1
                break
        return max_col

    def get_maximum_rows(sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows

    def mark_present(self,date,name,intime,outime):
        rows = self.get_maximum_rows(self.m1)
        self.m1.cell(rows+1,1).value= date
        self.m1.cell(rows+1,2).value= name
        self.m1.cell(rows+1,3).value= intime
        self.m1.cell(rows+1,4).value= outime
        self.m1.cell(rows+1,6).value= "P"
        self.wb.save(path)