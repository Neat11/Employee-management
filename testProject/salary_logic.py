from openpyxl import load_workbook
from datetime import datetime
from destination import path
wb = load_workbook(path)
mnthsum = wb["MNTHSUM1"]
m1 = wb["M1"]
mast1 = wb["MAST1"]

class findDimensions:
    def get_maximum_cols(sheet):
        max_col = 0
        for i in range(1, sheet.max_column+2):
            if (sheet.cell(row=2, column= i).value == None) and (sheet.cell(row=1, column= i).value == None):
                max_col = i-1
                break
        return max_col

    def get_maximum_rows(sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows

class Salary_calculation:
    def calculate(values,mast1_headings,LDOA, DOA, no_of_days):
        global m1,mnthsum,wb,mast1
        mast1_rows = findDimensions.get_maximum_rows(mast1)+1
        mnthsum_rows = findDimensions.get_maximum_rows(mnthsum)+1
        m1_rows = findDimensions.get_maximum_rows(m1)+1
        mnthsum_cols = findDimensions.get_maximum_cols(mnthsum)
        tfcodes,names,leave_bal,lallow,bpay,t_all,sw_all,hra,o_all,m_all,adv_ded,comp,adv_bal,ptaxes,amnt_dpst,mnth_dpst,deposit_collected,m1_data,days_not_present,lvinit=[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
        for i in range(2,mast1_rows):
            if (mast1.cell(row=i,column=mast1_headings.index('STATUS')+1).value=="Active"):
                tfcodes.append(mast1.cell(row=i,column=mast1_headings.index('TFCODE')+1).value)
                names.append(mast1.cell(row=i,column=mast1_headings.index('NAME')+1).value)
                leave_bal.append(int(mast1.cell(row=i,column=mast1_headings.index('LVMEBAL')+1).value))
                lvinit.append(mast1.cell(row=i,column=mast1_headings.index('LEAVE_INIT')+1).value)
                lallow.append(int(mast1.cell(row=i,column=mast1_headings.index('LALLOW')+1).value))
                bpay.append(int(mast1.cell(row=i,column=mast1_headings.index('BPAY')+1).value))
                t_all.append(int(mast1.cell(row=i,column=mast1_headings.index('T_ALL')+1).value))
                sw_all.append(int(mast1.cell(row=i,column=mast1_headings.index('SW_ALL')+1).value))
                hra.append(int(mast1.cell(row=i,column=mast1_headings.index('HRA')+1).value))
                o_all.append(int(mast1.cell(row=i,column=mast1_headings.index('O_ALL')+1).value))
                m_all.append(int(mast1.cell(row=i,column=mast1_headings.index('M_ALL')+1).value))
                adv_ded.append(int(mast1.cell(row=i,column=mast1_headings.index('ADV_DEDRT')+1).value))
                comp.append(mast1.cell(row=i,column=mast1_headings.index('COMP')+1).value)
                adv_bal.append(int(mast1.cell(row=i,column=mast1_headings.index('ADV_BAL')+1).value))
                ptaxes.append(mast1.cell(row=i,column=mast1_headings.index('PTAXES')+1).value)
                amnt_dpst.append(int(mast1.cell(row=i,column=mast1_headings.index('AMNT_DPST')+1).value))
                mnth_dpst.append(int(mast1.cell(row=i,column=mast1_headings.index('MNTH_DPST')+1).value))
                deposit_collected.append(int(mast1.cell(row=i,column=mast1_headings.index('DPST_COLL')+1).value))
        percentage = [1 for i in range(len(names))]
        absent = [0 for i in range(len(names))]
        for i in tfcodes:
            x = values[i].split(",")
            dates = []
            if x[0] != '':
                dates = [datetime.strptime(j.split("-")[0].strip(), "%d.%m.%Y") for j in x]
                reason = [j.split("-")[1].strip() for j in x]
                lvmbal = leave_bal[tfcodes.index(i)]
                lv_init = lvinit[tfcodes.index(i)]
                if lv_init.month == 12:
                    lv_init = lv_init.replace(year=LDOA.year)
                else:
                    lv_init = lv_init.replace(year=DOA.year)
                LA =[]
                for j in range(len(dates)):
                    if LDOA<lv_init<DOA:
                        leave_bal[tfcodes.index(i)] = lallow[tfcodes.index(i)]
                        if dates[j] < lv_init:
                            if lvmbal>0:
                                LA.append("L")
                                lvmbal-=1
                            else:
                                LA.append("A")
                                absent[tfcodes.index(i)]+=1
                        elif dates[j] >= lv_init:
                            LA.append("L")
                            leave_bal[tfcodes.index(i)]-=1
                    else:
                        if lvmbal>0:
                            LA.append("L")
                            leave_bal[tfcodes.index(i)]-=1
                            lvmbal-=1
                        else:
                            LA.append("A")
                            absent[tfcodes.index(i)]+=1
                m1_data.append([[i.date() for i in dates], [i for j in range(len(dates))], reason,LA]) #.date() added
            days_not_present.append(len(dates))
        advance_deducted = [adv_ded[i] if adv_ded[i]<adv_bal[i] else adv_bal[i] for i in range(len(names))]
        adv_bal = [adv_bal[i]-advance_deducted[i] for i in range(len(names))]
        adv_ded = [adv_ded[i] if adv_bal[i]>adv_ded[i] else adv_bal[i] for i in range(len(names))]
        deposit = [amnt_dpst[i] if mnth_dpst[i]>0 else 0 for i in range(len(names))]
        deposit_collected = [deposit_collected[i]+deposit[i] for i in range(len(names))]
        mnth_dpst = [mnth_dpst[i]-1 if mnth_dpst[i]>0 else 0 for i in range(len(names))]
        salary = {}
        lavail = [lallow[i]-leave_bal[i] for i in range(len(names))]
        for i in range(len(names)):
            salary[names[i]]= bpay[i]+t_all[i]+sw_all[i]+hra[i]+o_all[i]+m_all[i]
            print(names[i],bpay[i],t_all[i],sw_all[i],hra[i],o_all[i],m_all[i])
            ideal_salary = salary[names[i]]
            salary[names[i]]= salary[names[i]]-advance_deducted[i]-((absent[i]*salary[names[i]]/no_of_days)*100//1)/100
            percentage[i]=salary[names[i]]/ideal_salary
        ptaxs = []
        for i in range(len(names)):
            if (ptaxes[i] == True) or (ptaxes[i] == "TRUE"):
                x = salary[names[i]]
                if 10000<x<15001:
                    ptaxs.append(110)
                elif 15000<x<25001:
                    ptaxs.append(130)
                elif 25000<x<40001:
                    ptaxs.append(150)
                elif 40000<x:
                    ptaxs.append(200)
                else:
                    ptaxs.append(0)
            else:
                ptaxs.append(0)
        dd = [tfcodes, names, [DOA.date() for i in range(len(names))],absent,[days_not_present[i]-absent[i] for i in range(len(names))],[no_of_days-absent[i] for i in range(len(names))], [bpay[i]*percentage[i]*100//1/100 for i in range(len(names))],[t_all[i]*percentage[i]*100//1/100 for i in range(len(names))],[sw_all[i]*percentage[i]*100//1/100 for i in range(len(names))],[hra[i]*percentage[i]*100//1/100 for i in range(len(names))],[m_all[i]*percentage[i]*100//1/100 for i in range(len(names))],[o_all[i]*percentage[i]*100//1/100 for i in range(len(names))],adv_ded,[salary[names[i]]-ptaxs[i] for i in range(len(names))],leave_bal,comp,ptaxs,deposit]
        
        for i in range(1,mnthsum_cols+1):
            for j in range(len(names)):
                mnthsum[chr(ord('A')+i-1)+str(mnthsum_rows+j)] = dd[i-1][j]
        i=0
        for j in range(mast1_rows-2):
            if (mast1.cell(row=j+2,column=mast1_headings.index('STATUS')+1).value=="Active"):
                mast1.cell(row = j+2, column=mast1_headings.index('LVMEBAL')+1).value = leave_bal[i]
                mast1.cell(row = j+2, column=mast1_headings.index('LAVAIL')+1).value = lavail[i]
                mast1.cell(row = j+2, column=mast1_headings.index('ADV_BAL')+1).value = adv_bal[i]
                mast1.cell(row = j+2, column=mast1_headings.index('ADV_DEDRT')+1).value = adv_ded[i]
                mast1.cell(row = j+2, column=mast1_headings.index('MNTH_DPST')+1).value = mnth_dpst[i]
                mast1.cell(row = j+2, column=mast1_headings.index('DPST_COLL')+1).value = deposit_collected[i]
                i+=1
        for cell in mast1["C"]:
            cell.number_format = "DD-MM-YYYY"
        for cell in mnthsum["C"]:
            cell.number_format = "DD-MM-YYYY"  #added
        for cell in m1["A"]:
            cell.number_format = "DD-MM-YYYY"  #added
        for i in m1_data:
            for j in range(len(i[0])):
                m1.cell(row=m1_rows+j,column=1).value=i[0][j]
                m1.cell(row=m1_rows+j,column=2).value=i[1][j]
                m1.cell(row=m1_rows+j,column=5).value=i[2][j]
                m1.cell(row=m1_rows+j,column=6).value=i[3][j]
            m1_rows+=len(i[0])
        wb.save(path)
        return salary
