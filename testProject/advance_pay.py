from openpyxl import load_workbook
from destination import path

wb = load_workbook(path)
sheet = wb["MAST1"]

def get_maximum_rows(sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows

def get_maximum_cols(sheet):
        max_col = 0
        for i in range(1, sheet.max_column+2):
            if (sheet.cell(row=2, column= i).value == None) and (sheet.cell(row=1, column= i).value == None):
                max_col = i
                break
        return max_col
    
rows = get_maximum_rows(sheet)
cols = get_maximum_cols(sheet)

def add_advance(tfcode,name,amount,months,keylist):
    for i in range(1,rows+1):
        if sheet[chr(ord("A")+keylist.index("TFCODE"))+str(i)].value == tfcode:
            if sheet[chr(ord("A")+keylist.index("NAME"))+str(i)].value == name:
                sheet[chr(ord("A")+keylist.index("ADV_BAL"))+str(i)] = amount
                if (months != "-"):  
                    sheet[chr(ord("A")+keylist.index("ADV_DEDRT"))+str(i)] = int(amount)/int(months)
                    wb.save(path)
                    return "Advance Added"
                else:
                    return "Month cannot be left empty"
            else:
                return "Name does not match Tfcode"
    else:
        return "No tfcode found"
